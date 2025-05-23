import openpyxl
from collections import defaultdict
import os
from datetime import datetime

# Skriptpfad & Excel-Datei
skript_ordner = os.path.dirname(os.path.abspath(__file__))
excel_datei = "Dienstpläne2025.xlsx"
DATEIPFAD = os.path.join(skript_ordner, excel_datei)

# Workbook laden
wb = openpyxl.load_workbook(DATEIPFAD)
SP_WOCHENTAG = 0  # Spalte A
SP_DATUM = 1      # Spalte B
SP_VORDERGRUND = 3  # Spalte D
SP_RUFDIENST = 4    # Spalte E

# Punktesystem
PUNKTE = {"ruf": 0.5, "vouw": 1, "vowo": 2}
WOCHENENDE_TAGE = {4, 5, 6}  # Freitag, Samstag, Sonntag (0=Montag)

# Farbcodes für eingesprungene Dienste (rot, blau)
ERLAUBTE_FARBEN = {"FFFF0000", "FF0000FF"}

# Datenspeicher
daten = defaultdict(lambda: {"ruf": 0, "vouw": 0, "vowo": 0})

def ist_eingesprungen(zelle):
    if not zelle or not zelle.font or not zelle.font.color:
        return False
    color = zelle.font.color
    return color.type == "rgb" and color.rgb.upper() in ERLAUBTE_FARBEN

def ist_gueltiger_name(name):
    if not name:
        return False
    name = str(name).strip().lower()
    return name not in {"rufdienst", "vordergrund", "gesamt", "vowo", "vouw", ""}

def get_wochentag(row):
    """Wochentag aus Spalte A (Name) oder Spalte B (Datum) berechnen"""
    wert = row[SP_WOCHENTAG].value
    if isinstance(wert, str):
        name = wert.strip().lower()
        mapping = {"montag": 0, "dienstag": 1, "mittwoch": 2, "donnerstag": 3,
                   "freitag": 4, "samstag": 5, "sonntag": 6, "mo": 0, "di": 1, "mi": 2,
                   "do": 3, "fr": 4, "sa": 5, "so": 6}
        return mapping.get(name, None)
    elif isinstance(row[SP_DATUM].value, datetime):
        return row[SP_DATUM].value.weekday()
    return None

# Dienstblätter durchgehen
for sheetname in wb.sheetnames:
    if sheetname == "Einspringer-Statistik":
        continue
    ws = wb[sheetname]
    for row in ws.iter_rows(min_row=2):
        wotag = get_wochentag(row)
        print(wotag)
        ist_we = wotag in WOCHENENDE_TAGE if wotag is not None else False

        # Vordergrunddienst prüfen
        zelle_vg = row[SP_VORDERGRUND]
        name_vg = str(zelle_vg.value).strip() if zelle_vg.value else ""
        if ist_gueltiger_name(name_vg) and ist_eingesprungen(zelle_vg):
            if ist_we:
                daten[name_vg]["vowo"] += 1
            else:
                daten[name_vg]["vouw"] += 1

        # Rufdienst prüfen
        zelle_ruf = row[SP_RUFDIENST]
        name_ruf = str(zelle_ruf.value).strip() if zelle_ruf.value else ""
        if ist_gueltiger_name(name_ruf) and ist_eingesprungen(zelle_ruf):
            daten[name_ruf]["ruf"] += 1

# Statistik-Blatt neu anlegen
if "Einspringer-Statistik" in wb.sheetnames:
    del wb["Einspringer-Statistik"]
ws_stat = wb.create_sheet("Einspringer-Statistik")
ws_stat.append(["Name", "Ruf", "VoUW", "VoWE", "Gesamt"])

# Statistik eintragen, sortiert nach Gesamtpunkten
for name, werte in sorted(daten.items(), key=lambda x: (
    x[1]["ruf"] * PUNKTE["ruf"] +
    x[1]["vouw"] * PUNKTE["vouw"] +
    x[1]["vowo"] * PUNKTE["vowo"]
), reverse=True):
    gesamt = (
        werte["ruf"] * PUNKTE["ruf"] +
        werte["vouw"] * PUNKTE["vouw"] +
        werte["vowo"] * PUNKTE["vowo"]
    )
    ws_stat.append([name, werte["ruf"], werte["vouw"], werte["vowo"], round(gesamt, 1)])

wb.save(DATEIPFAD)
print("✅ Einspringer-Statistik erfolgreich gespeichert.")
