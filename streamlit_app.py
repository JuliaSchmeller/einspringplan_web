import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
from Einspringprogramm import berechne_einspringer_statistik

st.set_page_config(page_title="Einspringprogramm", layout="centered")
st.title("📊 Einspringprogramm Web-App")

uploaded_file = st.file_uploader(
    "📎 Excel-Datei hochladen (z. B. Dienstpläne2025.xlsx)", type=["xlsx"]
)

if uploaded_file is not None:
    try:
        wb = load_workbook(uploaded_file)
        wb = berechne_einspringer_statistik(wb)

        st.success("✅ Statistik erfolgreich berechnet!")

        # Vorschau (Name in A2 zeigen)
        ws_stat = wb["Einspringer-Statistik"]
        if ws_stat.max_row >= 2:
            beispielname = ws_stat["A2"].value
            st.info(f"👤 Erster Name in Statistik: **{beispielname}**")
        else:
            st.warning("⚠️ Keine Namen in der Statistik gefunden.")

        # In BytesIO speichern
        excel_bytes = BytesIO()
        wb.save(excel_bytes)
        excel_bytes.seek(0)

        # Download-Button
        st.download_button(
            label="⬇️ Bearbeitete Datei herunterladen",
            data=excel_bytes,
            file_name="Dienstpläne2025_mit_Statistik.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        st.error(f"❌ Fehler beim Verarbeiten der Datei: {e}")
